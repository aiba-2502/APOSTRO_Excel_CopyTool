from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from config import config  
from logger import logger  


def load_workbook_with_logging(file_path, mode="input"):
    """ファイルを読み込み、エラーログを出力"""
    try:
        logger.info(f"{mode.capitalize()}ファイル '{file_path}' を読み込みます")
        return load_workbook(file_path)
    except FileNotFoundError:
        logger.error(f"{mode.capitalize()}ファイル '{file_path}' が見つかりません")
        raise


def copy_template_sheet(output_wb, template_sheet_name, output_sheet_name):
    """テンプレートシートを複製し、フィルター設定も複製する"""
    if template_sheet_name not in output_wb.sheetnames:
        logger.error(f"テンプレートシート '{template_sheet_name}' が存在しません")
        raise ValueError(f"テンプレートシート '{template_sheet_name}' が出力ファイルに存在しません。")

    if output_sheet_name in output_wb.sheetnames:
        logger.info(f"既存のシート '{output_sheet_name}' を削除します")
        output_wb.remove(output_wb[output_sheet_name])

    logger.info(f"テンプレートシート '{template_sheet_name}' を複製します")
    template_sheet = output_wb[template_sheet_name]
    new_sheet = output_wb.copy_worksheet(template_sheet)
    new_sheet.title = output_sheet_name

    # オートフィルターの複製
    if template_sheet.auto_filter.ref:
        logger.info("フィルター設定を複製します")
        new_sheet.auto_filter.ref = template_sheet.auto_filter.ref

        for filter_column in template_sheet.auto_filter.filterColumn:
            col_id = filter_column.colId
            vals = filter_column.vals if hasattr(filter_column, 'vals') else []
            new_sheet.auto_filter.add_filter_column(col_id, vals=vals)

    return new_sheet


def transfer_values(value_sheet, output_sheet, copy_range, paste_start):
    """値をコピーして貼り付け"""
    logger.info("値の転記を開始します")
    min_col, min_row, max_col, max_row = range_boundaries(copy_range)
    paste_col, paste_row = range_boundaries(f"{paste_start}:{paste_start}")[:2]

    for i, row in enumerate(value_sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)):
        for j, cell in enumerate(row):
            output_cell = output_sheet.cell(row=paste_row + i, column=paste_col + j)
            output_cell.value = cell.value

    # 貼り付け範囲の行数を返す
    return max_row - min_row + 1


def delete_below_rows(output_sheet, paste_row, max_row_in_paste):
    """貼り付け範囲より下の行を削除"""
    logger.info("貼り付け範囲より下の行を削除します")

    # 貼り付け範囲の終了行
    last_paste_row = paste_row + max_row_in_paste - 1

    # 貼り付け範囲の終了行の次から削除
    if last_paste_row < output_sheet.max_row:
        rows_to_delete = output_sheet.max_row - last_paste_row
        output_sheet.delete_rows(last_paste_row + 1, rows_to_delete)
        logger.info(f"行 {last_paste_row + 1} 以降を削除しました")


def adjust_row_heights(output_sheet, copy_range, paste_start, min_row_height, default_font_size, line_height_multiplier):
    """貼り付け範囲の行の高さを自動調整"""
    logger.info("行の高さを調整します")
    min_col, min_row, max_col, max_row = range_boundaries(copy_range)
    paste_col, paste_row = range_boundaries(f"{paste_start}:{paste_start}")[:2]

    for row_index in range(paste_row, paste_row + (max_row - min_row) + 1):
        max_height = min_row_height
        for col_index in range(paste_col, paste_col + (max_col - min_col) + 1):
            cell = output_sheet.cell(row=row_index, column=col_index)
            if cell.value is not None:
                cell_lines = str(cell.value).split("\n")
                line_count = len(cell_lines)
                line_height = default_font_size * line_height_multiplier
                calculated_height = line_height * line_count
                max_height = max(max_height, calculated_height)
        output_sheet.row_dimensions[row_index].height = max_height


def hide_gridlines(sheet):
    """目盛り線（グリッドライン）を非表示にする"""
    try:
        logger.info("目盛り線を非表示に設定します")
        sheet.sheet_view.showGridLines = False
    except Exception as e:
        logger.error(f"目盛り線の非表示設定中にエラーが発生しました: {e}")
        raise


def save_workbook(workbook, file_path):
    """ワークブックを保存する"""
    try:
        logger.info(f"出力ファイル '{file_path}' に保存します")
        workbook.save(file_path)
    except Exception as e:
        logger.error(f"ワークブックの保存中にエラーが発生しました: {e}")
        raise


def main():
    files = config["files"]
    sheets = config["sheets"]
    copy_settings = config["copy_settings"]
    row_height_settings = config["row_height_settings"]
    value_file = files["value_file"]
    output_file = files["output_file"]
    template_sheet_name = sheets["template_sheet_name"]
    value_sheet_name = sheets["value_sheet_name"]
    output_sheet_name = sheets["output_sheet_name"]
    copy_range = copy_settings["copy_range"]
    paste_start = copy_settings["paste_start"]
    default_font_size = row_height_settings["default_font_size"]
    min_row_height = row_height_settings["min_row_height"]
    line_height_multiplier = row_height_settings["line_height_multiplier"]
    

    try:
        value_wb = load_workbook_with_logging(value_file, mode="値")
        value_sheet = value_wb[value_sheet_name]

        output_wb = load_workbook_with_logging(output_file, mode="出力")
        output_sheet = copy_template_sheet(output_wb, template_sheet_name, output_sheet_name)

        max_row_in_paste = transfer_values(value_sheet, output_sheet, copy_range, paste_start)

        delete_below_rows(output_sheet, range_boundaries(paste_start)[1], max_row_in_paste)

        adjust_row_heights(output_sheet, copy_range, paste_start, min_row_height, default_font_size, line_height_multiplier)

        hide_gridlines(output_sheet)  # グリッドライン非表示関数の呼び出し
        
        save_workbook(output_wb, output_file)  # ワークブック保存関数の呼び出し

        logger.info("スクリプトの実行が完了しました")
        logger.info(" ")

    except Exception as e:
        logger.error(f"エラーが発生しました: {e}")
        raise


if __name__ == "__main__":
    main()
